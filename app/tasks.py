import os
from celery import shared_task
from openpyxl import load_workbook
from app.models import Cuisine, Recipe, RecipeIngredient
from users.models import CustomUser as User


@shared_task
def process_bulk_recipes(file_path, user_id):
    user = User.objects.get(id=user_id)

    if not os.path.exists(file_path):
        return

    wb = load_workbook(file_path)
    sheet = wb.active
    rows = list(sheet.iter_rows(min_row=2, values_only=True))

    chunk_size = 100
    for i in range(0, len(rows), chunk_size):
        chunk = rows[i:i+chunk_size]
        process_chunk(chunk, user)

    os.remove(file_path)


def process_chunk(chunk, user):
    for row in chunk:
        try:
            title, description, instructions, prep_duration, cook_duration, cuisine_id, ingredient_ids_str = row
            title = title.strip()
            description = description.strip()
            instructions = instructions.strip()
            cuisine = Cuisine.objects.get(pk=cuisine_id)

            recipe = Recipe.objects.create(
                creator=user,
                title=title,
                description=description,
                instructions=instructions,
                prep_duration=int(prep_duration),
                cook_duration=int(cook_duration),
                cuisine=cuisine
            )

            ingredient_ids = [id.strip() for id in ingredient_ids_str.split(",")]
            for ing_id in ingredient_ids:
                if ing_id:
                    RecipeIngredient.objects.create(recipe=recipe, ingredient_id=ing_id)

            print(f"✅ Created recipe: {recipe.title}")

        except Exception as e:
            print(f"❌ Failed to process row: {row} -> Error: {e}")
            continue
